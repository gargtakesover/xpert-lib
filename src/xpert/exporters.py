"""
Export module for Xpert.
Supports CSV, Excel, JSON, and Markdown formats.
"""

import csv
import io
import json
import os
from typing import List, Optional, Union


TWEET_COLUMNS = [
    "id", "url", "author", "author_display", "text",
    "created_at", "likes", "retweets", "replies", "views",
    "content_type", "is_reply", "is_retweet", "is_pinned",
    "is_thread", "thread_position", "thread_length",
]


def _clean_dict(d: dict, full_data: bool) -> dict:
    """Remove empty/falsy fields unless full_data is requested."""
    if full_data:
        return d
    return {k: v for k, v in d.items() if v not in (None, "", [], {})}


def _flatten_tweet(tweet, full_data: bool = False) -> dict:
    """Flatten a Tweet object into a dict for CSV/Excel."""
    d = {
        "id": tweet.id,
        "url": tweet.url,
        "author": tweet.author,
        "author_display": tweet.author_display,
        "text": tweet.text,
        "created_at": tweet.created_at,
        "likes": tweet.likes,
        "retweets": tweet.retweets,
        "replies": tweet.replies,
        "views": tweet.views,
        "content_type": tweet.content_type,
        "is_reply": tweet.is_reply,
        "is_retweet": tweet.is_retweet,
        "is_pinned": tweet.is_pinned,
        "is_thread": tweet.is_thread,
        "thread_position": tweet.thread_position,
        "thread_length": tweet.thread_length,
    }
    return _clean_dict(d, full_data)


def tweets_to_csv(tweets: List, output: Optional[str] = None, full_data: bool = False) -> str:
    """Export tweets to CSV."""
    rows = [_flatten_tweet(t, full_data) for t in tweets]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=TWEET_COLUMNS)
    writer.writeheader()
    writer.writerows(rows)
    content = buf.getvalue()
    if output:
        with open(output, "w", newline="", encoding="utf-8") as f:
            f.write(content)
        return output
    return content


def tweets_to_json(tweets: List, output: Optional[str] = None, pretty: bool = True, full_data: bool = False) -> str:
    """Export tweets to JSON."""
    data = [_flatten_tweet(t, full_data) for t in tweets]
    indent = 2 if pretty else None
    content = json.dumps(data, indent=indent, ensure_ascii=False)
    if output:
        with open(output, "w", encoding="utf-8") as f:
            f.write(content)
        return output
    return content


def tweets_to_markdown(tweets: List, output: Optional[str] = None, full_data: bool = False) -> str:
    """Export tweets to Markdown table."""
    lines = ["# Xpert Scrape Results\n"]
    lines.append("| # | Author | Text | Likes | Retweets | Date |")
    lines.append("|---|--------|------|-------|----------|------|")
    for i, t in enumerate(tweets, 1):
        text = t.text.replace("|", "\\|").replace("\n", "<br>")
        lines.append(f"| {i} | @{t.author} | {text} | {t.likes} | {t.retweets} | {t.created_at[:10]} |")
    content = "\n".join(lines)
    if output:
        with open(output, "w", encoding="utf-8") as f:
            f.write(content)
        return output
    return content


def tweets_to_excel(tweets: List, output: Optional[str] = None, full_data: bool = False) -> Union[bytes, str]:
    """Export tweets to Excel (.xlsx)."""
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandas and openpyxl required for Excel export. Install with: pip install xpert[excel]")

    rows = [_flatten_tweet(t, full_data) for t in tweets]
    df = pd.DataFrame(rows, columns=TWEET_COLUMNS)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tweets")
        ws = writer.sheets["Tweets"]
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1DA1F2", end_color="1DA1F2", fill_type="solid")
        for col_idx in range(1, len(TWEET_COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    data = buf.getvalue()
    if output:
        with open(output, "wb") as f:
            f.write(data)
        return output
    return data


def tweets_to_format(tweets: List, fmt: str, output: str, full_data: bool = False) -> str:
    """Export tweets in specified format. Writes to output file."""
    fmt = fmt.lower()
    if fmt == "csv":
        return tweets_to_csv(tweets, output, full_data)
    elif fmt == "excel":
        return tweets_to_excel(tweets, output, full_data)
    elif fmt == "markdown":
        return tweets_to_markdown(tweets, output, full_data)
    elif fmt == "json":
        return tweets_to_json(tweets, output, full_data=full_data)
    else:
        raise ValueError(f"Unknown format: {fmt}. Use: csv, excel, json, markdown")


USER_COLUMNS = ["username", "display_name", "bio", "followers", "following", "tweets", "url", "profile_picture", "joined", "verified", "location", "website"]

def _flatten_user(user, full_data: bool = False) -> dict:
    """Flatten a User object for export."""
    d = {
        "username": user.username,
        "display_name": user.display_name,
        "bio": user.bio,
        "followers": user.followers,
        "following": user.following,
        "tweets": user.tweets,
        "url": user.url,
        "profile_picture": user.profile_picture,
        "joined": user.joined,
        "verified": user.verified,
        "location": getattr(user, "location", ""),
        "website": getattr(user, "website", ""),
    }
    return _clean_dict(d, full_data)


def users_to_csv(users: List, output: Optional[str] = None, full_data: bool = False) -> str:
    """Export users to CSV."""
    rows = [_flatten_user(u, full_data) for u in users]
    if not rows:
        rows = [{}]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=USER_COLUMNS)
    writer.writeheader()
    writer.writerows(rows)
    content = buf.getvalue()
    if output:
        with open(output, "w", newline="", encoding="utf-8") as f:
            f.write(content)
        return output
    return content


def users_to_excel(users: List, output: Optional[str] = None, full_data: bool = False) -> Union[bytes, str]:
    """Export users to Excel (.xlsx)."""
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandas and openpyxl required for Excel export. Install with: pip install xpert[excel]")

    rows = [_flatten_user(u, full_data) for u in users]
    if not rows:
        rows = [dict.fromkeys(USER_COLUMNS, "")]
    df = pd.DataFrame(rows, columns=USER_COLUMNS)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Profiles")
    data = buf.getvalue()
    if output:
        with open(output, "wb") as f:
            f.write(data)
        return output
    return data


def users_to_markdown(users: List, output: Optional[str] = None, full_data: bool = False) -> str:
    """Export users to Markdown."""
    lines = ["# Xpert User Profiles\n"]
    lines.append("| Username | Display Name | Followers | Following | Verified |")
    lines.append("|----------|--------------|-----------|----------|---------|")
    for u in users:
        d = _flatten_user(u, full_data)
        verified = "Yes" if d.get("verified") else ""
        lines.append(f"| @{d.get('username', '')} | {d.get('display_name', '')} | {d.get('followers', 0):,} | {d.get('following', 0):,} | {verified} |")
    content = "\n".join(lines)
    if output:
        with open(output, "w", encoding="utf-8") as f:
            f.write(content)
        return output
    return content


def users_to_format(users: List, fmt: str, output: str, full_data: bool = False) -> str:
    """Export users in specified format. Writes to output file."""
    fmt = fmt.lower()
    if fmt == "csv":
        return users_to_csv(users, output, full_data)
    elif fmt == "excel":
        return users_to_excel(users, output, full_data)
    elif fmt == "markdown":
        return users_to_markdown(users, output, full_data)
    elif fmt == "json":
        return json.dumps([_flatten_user(u, full_data) for u in users], indent=2, ensure_ascii=False)
    else:
        raise ValueError(f"Unknown format: {fmt}. Use: csv, excel, json, markdown")
